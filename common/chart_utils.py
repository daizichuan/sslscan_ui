from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

from SSLScan_UI.common.sqlite_utils import SqliteUtils


def time_comsuming_lineChart(website, sheet_name: str, excel_file):
    # 处理特殊字符
    special_chars = "[:：/\\]"
    for special_char in special_chars:
        sheet_name = sheet_name.replace(special_char, '_')

    lst = []
    lst.append(['timestamp', website])

    sqlite_res = SqliteUtils().get_time_consuming(website)
    for _ in sqlite_res:
        lst_tmp = []
        lst_tmp.append(str(_[2]))
        lst_tmp.append(_[1])
        lst.append(lst_tmp)

    wb = load_workbook(excel_file)

    try:
        sheet = wb[sheet_name]
    except KeyError:
        wb.create_sheet(sheet_name)
        sheet = wb[sheet_name]

    for row in lst:
        sheet.append(row)

    chart = LineChart()
    # 图的标题
    chart.title = "SSL Scan站点扫描耗时统计"
    # y轴标题
    chart.y_axis.title = "耗时时间（毫秒）"
    # x轴标题
    chart.x_axis.title = "时间戳"
    # 数据来源
    data = Reference(worksheet=sheet, min_col=2, min_row=1, max_col=2, max_row=sheet.max_row)
    # 设定x轴项目名称
    categories = Reference(sheet, min_col=1, min_row=2, max_col=1, max_row=sheet.max_row)
    # 折线图数据，y轴名称为第一行列名称
    # from_rows：如果是True表示将一行的数据作为一个条线，如果是False表示按列画
    # titles_from_data：如果是True，表示每一组（根据from_rows确定是每一行还是每一列）数据的第一个作为title
    chart.add_data(data, from_rows=False, titles_from_data=True)
    chart.set_categories(categories)
    sheet.add_chart(chart, "D10")

    sheet = wb['首页']
    sheet.append([f'=HYPERLINK("#{sheet_name}！A1","跳转{sheet_name}页")'])

    wb.save(excel_file)


if __name__ == '__main__':
    time_comsuming_lineChart('www.weiyoutong.cn', 'Sheet5', '../test/first.xlsx')
