import shutil

import pandas as pd
from tkinter import messagebox
from datetime import datetime
from openpyxl import load_workbook

from SSLScan_UI.common.log import log
from SSLScan_UI.common.diskcache_utils import DiskcacheUtils
from SSLScan_UI.common.sqlite_utils import SqliteUtils


class ExcelUtils:

    def __init__(self):
        self.test_data_excel = '../test_datas/SSL_Scan_data_all.xlsx'
        self.website_excel = '../test_datas/test_site.xlsx'

    def get_excel_service_url(self):
        df = pd.read_excel(self.test_data_excel).columns.values.tolist()[1:2]
        return ''.join(df)

    def get_excel_data(self, s_name='Sheet1'):
        return pd.read_excel(self.test_data_excel, sheet_name=s_name, header=1).to_dict(orient="records")

    def setback_excel_actual_results(self, results, s_name='Sheet1', f_name=''):

        df = pd.Series(results)

        try:
            writer = pd.ExcelWriter(f_name, engine='openpyxl', mode='a', if_sheet_exists="overlay")
            writer.sheets.update({sht.title: sht for sht in writer.book.worksheets})
            df.to_excel(writer, sheet_name=s_name, index=False, startrow=2, startcol=8, header=None)
            writer.close()
        except PermissionError:
            messagebox.showinfo(title='注意', message='Excel用例文件已被打开，请先关闭')
            raise KeyboardInterrupt(f'用例文件{f_name}被打开了')

    def setback_excel_accordingto_NO(self, f_name='', s_name=['Sheet1']):
        wb = load_workbook(f_name, data_only=True)
        for s_n in s_name:
            sheet = wb[s_n]
            for i, cell in enumerate(list(sheet.columns)[0]):
                if i > 1:
                    sheet[f'J{cell.row}'] = DiskcacheUtils().get_diskcache_results(cell.value)
                    sheet[f'k{cell.row}'] = f'=IF(I{cell.row}=J{cell.row},"通过","未通过")'
        wb.save(f_name)

    # 废弃，通过excel自带公式去算，做个模板
    def results_pass_or_not(self):
        data = pd.read_excel(self.test_data_excel, header=1)
        df = pd.DataFrame(data)
        df['success_or_failed'] = df.apply(lambda x: self._function(x['expected_result'], x['actual_result']), axis=1)

        # df.to_excel('../test_datas/SSL_Scan_data111.xlsx', index=False)

        try:
            writer = pd.ExcelWriter(self.test_data_excel, engine='openpyxl', mode='a', if_sheet_exists="overlay")
            writer.sheets.update({sht.title: sht for sht in writer.book.worksheets})
            df['success_or_failed'].to_excel(writer, sheet_name='Sheet1', index=False, startrow=2, startcol=9,
                                             header=None)
            writer.close()
        except PermissionError:
            messagebox.showinfo(title='注意', message='Excel用例文件已被打开，请先关闭')
            raise KeyboardInterrupt(f'用例文件{self.test_data_excel}被打开了')

    def _function(self, a, b):
        if a == b:
            return '成功'
        else:
            return '失败'

    def get_excel_website_url(self):
        df = pd.read_excel(self.website_excel).columns.values.tolist()[1]
        return df

    def get_excel_websites(self):
        # usecols=["B,C,E:G"] usecols=[1,2,4,5,6] usecols=["学号,姓名,学科,成绩,学年"] 列名 列索引 列名称都行
        df = pd.read_excel(self.website_excel, header=1, usecols=[0]).values.tolist()
        return df

    def setback_excel_websites_results(self, lst):
        file_name = f"../test_datas/test_site{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        log.info(f'生成结果文件{file_name}')
        shutil.copy(self.website_excel, file_name)

        try:
            writer = pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists="overlay")
            writer.sheets.update({sht.title: sht for sht in writer.book.worksheets})

            for _ in lst:
                df = pd.Series(_[0])
                df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=1, startcol=_[1], header=None)
                log.info(f'往结果文件{file_name}回写结果')

            writer.close()
        except PermissionError:
            messagebox.showinfo(title='注意', message='Excel用例文件已被打开，请先关闭')
            raise KeyboardInterrupt(f'用例文件{file_name}被打开了')

    def gen_websites_scan_excel(self, f_name='', res=''):
        wb = load_workbook(f_name)
        log.info(f'往{f_name}回写结果')
        sheet = wb['Sheet2']
        for _ in res:
            sheet.append(_)

        wb.save(f_name)


if __name__ == '__main__':
    # print(ExcelUtils().get_excel_websites())

    ExcelUtils().gen_websites_scan_excel('../test_datas/test_site20230510213803.xlsx',
                                         SqliteUtils().get_time_consuming_all())
